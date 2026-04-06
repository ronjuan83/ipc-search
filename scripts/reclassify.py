#!/usr/bin/env python3
"""
IPC 批次重分類標記工具

根據 WIPO IPC concordance 資料，標記專利的 IPC 代碼是否涉及版本異動，
並依處理難度分為三級：自動替換、需判斷（廢棄多目的地）、需確認（技術範圍移動）。

用法:
    python3 scripts/reclassify.py input.xlsx -c IPC
    python3 scripts/reclassify.py data.csv -c IPC -d 申請日 --verbose
    python3 scripts/reclassify.py input.xlsx -c IPC -d "申請日" -o output.xlsx
    python3 scripts/reclassify.py --self-test

版本推斷：
    若指定 -d 申請日欄位，系統會根據申請日推斷專利使用的 IPC 版本，
    只套用「在該版本之後發生的」異動。例如 2010 年申請的專利，
    只會標記 2010 年之後的異動，不會標記 1995→2000 年的異動。
"""

from __future__ import annotations
import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

# ── IPC 版本對照表 ───────────────────────────────────────────
# 申請日所在的年份 → 當時使用的 IPC 版本
# 參考：IPC 第 6 版 (1994)、第 7 版 (2000)、第 8 版/改革版 (2006.01 起每年更新)

IPC_VERSION_BY_YEAR = {
    # 1994-1999: IPC 第 6 版
    **{y: '1994.01' for y in range(1994, 2000)},
    # 2000-2005: IPC 第 7 版
    **{y: '2000.01' for y in range(2000, 2006)},
    # 2006 起：改革版，每年更新
    **{y: f'{y}.01' for y in range(2006, 2030)},
}


def year_to_ipc_version(year: int) -> str:
    """將年份轉為 IPC 版本字串。"""
    if year < 1994:
        return '1994.01'
    return IPC_VERSION_BY_YEAR.get(year, f'{year}.01')


def parse_apply_date(value) -> int | None:
    """解析申請日為西元年份。

    支援格式：
      20050101, 2005/01/01, 2005-01-01 (西元)
      0940101, 094/01/01 (民國)
      整數 20050101 或 940101
    """
    if not value:
        return None
    raw = str(value).strip().replace('/', '').replace('-', '').replace('.', '')
    # Remove non-digits
    raw = re.sub(r'\D', '', raw)
    if not raw:
        return None
    try:
        n = int(raw)
    except ValueError:
        return None

    if n > 19000000:
        # 西元 YYYYMMDD
        return n // 10000
    elif n > 800000:
        # 民國 YYYMMDD (e.g., 0940101 = 94年 = 2005)
        roc_year = n // 10000
        return roc_year + 1911
    elif n > 1900:
        # 純年份
        return n
    elif n > 80:
        # 民國年
        return n + 1911
    return None


# ── IPC 代碼解析 ─────────────────────────────────────────────

_IPC_RE = re.compile(r'^([A-H]\d{2}[A-Z])(?:\s*(\d+)/(\S+))?$')


def parse_ipc_code(code: str) -> dict | None:
    """拆解 IPC 代碼為 subclass / main_group / subgroup。

    支援多種格式：
    >>> parse_ipc_code('H01L 21/677')
    {'subclass': 'H01L', 'main': 21, 'sub': '677', 'group': 'H01L 21/677'}
    >>> parse_ipc_code('H01L21/677')     # 無空格
    {'subclass': 'H01L', 'main': 21, 'sub': '677', 'group': 'H01L 21/677'}
    >>> parse_ipc_code('H01L 21/677(2006.01)')  # 附版本標記
    {'subclass': 'H01L', 'main': 21, 'sub': '677', 'group': 'H01L 21/677'}
    >>> parse_ipc_code('H01L')
    {'subclass': 'H01L', 'main': None, 'sub': None, 'group': None}
    """
    code = code.strip().upper().replace('\n', ' ')
    code = re.sub(r'\s+', ' ', code)
    # Remove version/date suffix: (2006.01), [2006.01], (20060101)
    code = re.sub(r'[\(\[]\d{4}[\.\-]?\d{0,2}[\)\]]?$', '', code).strip()
    # Remove leading class indicator like "I:" or "N:" (TIPO format)
    code = re.sub(r'^[A-Z]:\s*', '', code)
    # Handle no-space format: "H01L21/677" → "H01L 21/677"
    m2 = re.match(r'^([A-H]\d{2}[A-Z])(\d+/\S+)$', code)
    if m2:
        code = f"{m2.group(1)} {m2.group(2)}"
    m = _IPC_RE.match(code)
    if not m:
        return None
    subclass, main, sub = m.group(1), m.group(2), m.group(3)
    if main is not None:
        return {
            'subclass': subclass,
            'main': int(main),
            'sub': sub,
            'group': f"{subclass} {main}/{sub}",
        }
    return {'subclass': subclass, 'main': None, 'sub': None, 'group': None}


def parse_ipc_field(value) -> list[str]:
    """將 IPC 欄位字串拆成清單。

    支援多種分隔格式：
      逗號: "H01L 21/677, G06F 3/01"
      分號: "H01L 21/677; G06F 3/01"
      換行: "H01L 21/677\\nG06F 3/01"
      空格+subclass: "H01L 21/677 G06F 3/01"（兩個代碼間無分隔符）
      TIPO pipe: "H01L 21/677|G06F 3/01"
    """
    if not value or (isinstance(value, float) and str(value) == 'nan'):
        return []
    raw = str(value).strip()
    if not raw:
        return []
    # Split by common delimiters
    parts = re.split(r'[,;|\n\r]+', raw)
    result = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        codes = re.findall(r'[A-H]\d{2}[A-Z]\s*\d+/\S+', part.upper())
        if len(codes) > 1:
            result.extend(c.strip() for c in codes)
        elif codes:
            result.append(codes[0].strip())
        else:
            sub = re.match(r'^[A-H]\d{2}[A-Z]$', part.strip().upper())
            if sub:
                result.append(sub.group())
    return result


# ── 核心分類器 ───────────────────────────────────────────────

class IPCReclassifier:
    """IPC 重分類標記引擎。

    三級處理：
      Level 1 (auto_replaced)    : 廢棄 subclass，單一目的地 → 自動替換
      Level 2 (review_deprecated): 廢棄 subclass，多個目的地 → 列出候選
      Level 3 (review_scope)     : group 級技術範圍移動 → 標記可能需移動
    """

    def __init__(self, data_path: str | Path | None = None):
        if data_path is None:
            data_path = Path(__file__).resolve().parent.parent / 'public' / 'ipc_data.json'
        with open(data_path, encoding='utf-8') as f:
            data = json.load(f)

        self.deprecated_to = data.get('deprecated_to', {})
        self.deprecated_at = data.get('deprecated_at', {})
        self.subclass_index = data.get('subclass_index', {})

        # 分類廢棄 subclass
        self.deprecated_single = {}
        self.deprecated_multi = {}
        for old, new in self.deprecated_to.items():
            if isinstance(new, list):
                self.deprecated_multi[old] = new
            else:
                self.deprecated_single[old] = new

        # 建立 group 級查找索引（保留所有版本記錄）
        self._build_all_donations()
        self._build_range_lookup()

        stats = (f"loaded: {len(self.deprecated_single)} auto-replace, "
                 f"{len(self.deprecated_multi)} multi-dest, "
                 f"{len(self.exact_lookup)} exact groups, "
                 f"{sum(len(v) for v in self.range_lookup.values())} ranges")
        self._stats = stats

    @staticmethod
    def _version_sort_key(version_str: str) -> str:
        """從 '1995.01→2000.01' 取得排序 key '2000.01'。"""
        if '→' in version_str:
            return version_str.split('→')[1].strip()
        return version_str

    @staticmethod
    def _version_from_key(version_str: str) -> str:
        """從 '1995.01→2000.01' 取得起始版本 '1995.01'。"""
        if '→' in version_str:
            return version_str.split('→')[0].strip()
        return version_str

    def _build_all_donations(self):
        """建立 src_group → 所有版本的 donated 記錄。"""
        all_donations = defaultdict(list)
        for sub, info in self.subclass_index.items():
            for d in info.get('donated', []):
                src = d.get('src_group', '').replace('\n', ' ').strip()
                dst = d.get('dst', '')
                ver = d.get('version', '')
                if src and dst:
                    all_donations[src].append({
                        'dst': dst,
                        'version': ver,
                        'src_sub': sub,
                        'ver_from': self._version_from_key(ver),
                        'ver_to': self._version_sort_key(ver),
                    })

        # 排序：最新版在前
        self.all_donations = {}
        for src, records in all_donations.items():
            records.sort(key=lambda x: x['ver_to'], reverse=True)
            self.all_donations[src] = records

        # 向後相容：exact_lookup 取最新版
        self.exact_lookup = {}
        for src, records in self.all_donations.items():
            r = records[0]
            self.exact_lookup[src] = {
                'dst': r['dst'], 'version': r['version'], 'src_sub': r['src_sub'],
            }

    def _build_range_lookup(self):
        """建立範圍查找表。"""
        self.range_lookup = defaultdict(list)
        for src, info in self.exact_lookup.items():
            if ' - ' not in src:
                continue
            parts = src.split(' - ')
            if len(parts) != 2:
                continue
            left, right = parts[0].strip(), parts[1].strip()
            parsed_left = parse_ipc_code(left)
            if not parsed_left or parsed_left['main'] is None:
                continue
            sub = parsed_left['subclass']
            right_match = re.match(r'(\d+)/(\S+)', right)
            if not right_match:
                continue
            self.range_lookup[sub].append({
                'start_main': parsed_left['main'], 'start_sub': parsed_left['sub'],
                'end_main': int(right_match.group(1)), 'end_sub': right_match.group(2),
                'dst': info['dst'], 'version': info['version'], 'src_range': src,
            })

    def _extract_first_dst(self, dst_str: str, src_sub: str) -> str:
        """從 dst 字串提取第一個非自引用的目的地代碼。"""
        parts = [p.strip() for p in dst_str.split(',')]
        for part in parts:
            code = part.split(' - ')[0].strip()
            if not code:
                continue
            parsed = parse_ipc_code(code)
            if parsed and parsed['subclass'] == src_sub:
                continue
            return code
        return parts[0].split(' - ')[0].strip() if parts else dst_str

    def _in_range(self, main: int, sub: str, r: dict) -> bool:
        if main < r['start_main'] or main > r['end_main']:
            return False
        if main == r['start_main'] and main == r['end_main']:
            return r['start_sub'] <= sub <= r['end_sub']
        if main == r['start_main']:
            return sub >= r['start_sub']
        if main == r['end_main']:
            return sub <= r['end_sub']
        return True

    def _is_after_version(self, version_str: str, patent_version: str | None) -> bool:
        """判斷異動是否發生在專利版本之後。

        version_str: '1995.01→2000.01' (異動的版本轉換)
        patent_version: '2005.01' (專利使用的版本)

        如果專利版本是 2005.01，而異動發生在 2006.01→2007.01，
        那這個異動是在專利之後發生的 → 需要標記。
        如果異動是 1995.01→2000.01，那在專利申請時已經完成 → 不需標記。
        """
        if not patent_version:
            return True  # 沒有版本資訊 → 標記所有異動
        ver_from = self._version_from_key(version_str)
        # 異動的起始版本 >= 專利版本 → 專利申請時這個異動還沒發生
        return ver_from >= patent_version

    def classify_code(self, code: str, patent_version: str | None = None) -> dict:
        """對單一 IPC 代碼進行重分類標記。

        Args:
            code: IPC 代碼字串
            patent_version: 專利使用的 IPC 版本（如 '2005.01'），用於過濾異動。
                           None 表示檢查所有版本的異動。
        """
        code = code.strip().upper().replace('\n', ' ')
        code = re.sub(r'\s+', ' ', code)
        parsed = parse_ipc_code(code)

        if not parsed:
            return {'action': 'unchanged', 'original': code, 'detail': '無法解析'}

        sub = parsed['subclass']
        result = {'action': 'unchanged', 'original': code, 'detail': ''}

        # ── Level 1: 廢棄 subclass，單一目的地 → 自動替換 ──
        if sub in self.deprecated_single:
            at = self.deprecated_at.get(sub, '?')
            # 版本檢查：廢棄是否發生在專利版本之後？
            if patent_version and at != '?' and at < patent_version:
                # 專利申請時這個 subclass 已經廢棄了 → 代碼本身可能有問題
                # 仍然標記為自動替換
                pass
            new_sub = self.deprecated_single[sub]
            if parsed['group']:
                new_code = parsed['group'].replace(sub, new_sub, 1)
            else:
                new_code = new_sub
            return {
                'action': 'auto_replaced',
                'original': code,
                'result': new_code,
                'detail': f"{sub} 廢棄→{new_sub} ({at})",
            }

        # ── Level 2: 廢棄 subclass，多目的地 → 需判斷 ──
        if sub in self.deprecated_multi:
            dests = self.deprecated_multi[sub]
            at = self.deprecated_at.get(sub, '?')
            return {
                'action': 'review_deprecated',
                'original': code,
                'candidates': ', '.join(dests),
                'detail': f"{sub} 廢棄→{'/'.join(dests)} ({at})，需依專利內容判斷",
            }

        # ── Level 3: group 級異動（版本感知） ──
        if parsed['group']:
            group_key = parsed['group']

            # 3a: 精確匹配 — 檢查所有版本的記錄
            if group_key in self.all_donations:
                for rec in self.all_donations[group_key]:
                    if self._is_after_version(rec['version'], patent_version):
                        dst = self._extract_first_dst(rec['dst'], rec['src_sub'])
                        ver_note = f" (專利版本:{patent_version})" if patent_version else ""
                        return {
                            'action': 'review_scope',
                            'original': code,
                            'possible_new': dst,
                            'detail': f"{group_key}→{dst} ({rec['version']}){ver_note}，需確認",
                        }
                # 所有異動都在專利版本之前 → 不需處理
                return result

            # 3b: 範圍匹配
            if sub in self.range_lookup and parsed['main'] is not None:
                for r in self.range_lookup[sub]:
                    if self._in_range(parsed['main'], parsed['sub'], r):
                        if self._is_after_version(r['version'], patent_version):
                            dst = self._extract_first_dst(r['dst'], sub)
                            return {
                                'action': 'review_scope',
                                'original': code,
                                'possible_new': dst,
                                'detail': f"{code} 在範圍 {r['src_range']} 內→{dst} ({r['version']})，需確認",
                            }

            # 3c: 主組匹配（xx/00）
            main_group = f"{sub} {parsed['main']}/00"
            if main_group in self.all_donations and main_group != group_key:
                for rec in self.all_donations[main_group]:
                    if self._is_after_version(rec['version'], patent_version):
                        dst = self._extract_first_dst(rec['dst'], rec['src_sub'])
                        return {
                            'action': 'review_scope',
                            'original': code,
                            'possible_new': dst,
                            'detail': f"主組 {main_group} 有異動→{dst} ({rec['version']})，此子組可能受影響",
                        }

        return result

    def classify_field(self, value, patent_version: str | None = None) -> list[dict]:
        """處理多碼 IPC 欄位，逐一分類。"""
        codes = parse_ipc_field(value)
        if not codes:
            return [{'action': 'unchanged', 'original': str(value), 'detail': '空值或無效'}]
        return [self.classify_code(c, patent_version) for c in codes]


# ── I/O ──────────────────────────────────────────────────────

def detect_format(filepath: Path) -> str:
    ext = filepath.suffix.lower()
    if ext == '.xlsx':
        return 'xlsx'
    if ext == '.tsv':
        return 'tsv'
    if ext == '.csv':
        return 'csv'
    with open(filepath, encoding='utf-8') as f:
        first_line = f.readline()
        if '\t' in first_line:
            return 'tsv'
    return 'csv'


def read_input(filepath: Path, ipc_col: str, fmt: str, date_col: str | None = None) -> tuple[list[dict], list[str]]:
    """讀入資料，返回 (rows, headers)。"""
    required_cols = [ipc_col]
    if date_col:
        required_cols.append(date_col)

    if fmt == 'xlsx':
        try:
            import openpyxl
        except ImportError:
            print("錯誤：需要 openpyxl 套件來讀取 Excel 檔案", file=sys.stderr)
            print("安裝：pip install openpyxl", file=sys.stderr)
            sys.exit(1)
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h or '') for h in next(rows_iter)]
        for col in required_cols:
            if col not in headers:
                print(f"錯誤：找不到欄位 '{col}'，可用欄位: {headers}", file=sys.stderr)
                sys.exit(1)
        rows = []
        for row in rows_iter:
            rows.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))
        wb.close()
        return rows, headers
    else:
        delimiter = '\t' if fmt == 'tsv' else ','
        with open(filepath, encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames or []
            for col in required_cols:
                if col not in headers:
                    print(f"錯誤：找不到欄位 '{col}'，可用欄位: {headers}", file=sys.stderr)
                    sys.exit(1)
            rows = list(reader)
        return rows, headers


def write_output(filepath: Path, rows: list[dict], headers: list[str], fmt: str):
    """寫出結果。"""
    new_cols = ['reclassify_action', 'reclassify_result', 'reclassify_candidates',
                'reclassify_detail', 'ipc_version_inferred']
    all_headers = headers + [c for c in new_cols if c not in headers]

    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(all_headers)
        for row in rows:
            ws.append([row.get(h, '') for h in all_headers])
        wb.save(filepath)
    else:
        delimiter = '\t' if fmt == 'tsv' else ','
        with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=all_headers, delimiter=delimiter, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)


def generate_summary(results_count: dict, total: int, version_stats: dict | None = None) -> str:
    """產生摘要報告。"""
    lines = [
        '=' * 55,
        'IPC 重分類標記報告',
        '=' * 55,
        f'總筆數: {total:,}',
        '',
    ]

    for action, label in [
        ('unchanged', '不需處理'),
        ('auto_replaced', '已自動替換（廢棄 subclass 單一目的地）'),
        ('review_deprecated', '需人工判斷（廢棄 subclass 多目的地）'),
        ('review_scope', '需確認（group 技術範圍移動，看專利內容）'),
        ('parse_error', '解析失敗'),
    ]:
        count = results_count.get(action, 0)
        pct = count / total * 100 if total > 0 else 0
        lines.append(f'  {action:<22} {count:>8,} ({pct:>5.1f}%) — {label}')

    if version_stats:
        lines.append('')
        lines.append('IPC 版本分布（依申請日推斷）：')
        for ver, count in sorted(version_stats.items()):
            lines.append(f'  {ver}: {count:>8,}')

    return '\n'.join(lines)


# ── 自測 ─────────────────────────────────────────────────────

def self_test(rc: IPCReclassifier):
    """內建驗證。"""
    print("  格式解析測試：")
    parse_tests = [
        ('H01L21/677', 'H01L', 21, '677'),
        ('H01L 21/677(2006.01)', 'H01L', 21, '677'),
        ('h01l 21/677', 'H01L', 21, '677'),
        ('I: H01L 21/677', 'H01L', 21, '677'),
    ]
    for raw, exp_sub, exp_main, exp_subg in parse_tests:
        p = parse_ipc_code(raw)
        ok = p and p['subclass'] == exp_sub and p['main'] == exp_main and p['sub'] == exp_subg
        print(f"    {'✅' if ok else '❌'} {raw:<28} → {p['group'] if p else 'FAIL'}")

    field_tests = [
        ('H01L 21/677, G06F 3/01', 2),
        ('H01L 21/677; G06F 3/01', 2),
        ('H01L 21/677|G06F 3/01', 2),
        ('H01L 21/677 G06F 3/01', 2),
    ]
    for raw, exp_count in field_tests:
        codes = parse_ipc_field(raw)
        ok = len(codes) == exp_count
        print(f"    {'✅' if ok else '❌'} field({raw[:30]}) → {len(codes)} codes")

    print("\n  申請日解析測試：")
    date_tests = [
        ('20050101', 2005), ('2005/01/01', 2005), ('2005-01-01', 2005),
        ('0940101', 2005), ('094/01/01', 2005),   # 民國 94 年
        (20100315, 2010), ('1120601', 2023),        # 民國 112 年
    ]
    for raw, exp_year in date_tests:
        year = parse_apply_date(raw)
        ok = year == exp_year
        print(f"    {'✅' if ok else '❌'} {str(raw):<16} → {year} (expect {exp_year})")

    print("\n  版本感知分類測試：")
    ver_tests = [
        # H01L 21/00 在 1995→2000 有異動。2010 年申請的專利不應標記此異動。
        ('H01L 21/00', '2010.01', None),   # 1995 異動已完成，但 2025 異動未完成 → 仍 review_scope
        # G06F 3/01 沒有異動
        ('G06F 3/01', '2010.01', 'unchanged'),
        # C13C 廢棄 — 無論版本都應標記
        ('C13C 3/02', '2020.01', 'auto_replaced'),
    ]
    for code, ver, exp_action in ver_tests:
        result = rc.classify_code(code, ver)
        if exp_action:
            ok = result['action'] == exp_action
        else:
            ok = True  # Just check it doesn't crash
        print(f"    {'✅' if ok else '❌'} {code:<16} ver={ver} → {result['action']}")

    print("\n  分類邏輯測試：")
    tests = [
        ('C13C 3/02', 'auto_replaced', 'result', 'C13B 3/02'),
        ('C13D', 'auto_replaced', 'result', 'C13B'),
        ('G06C 15/00', 'auto_replaced', 'result', 'G06N 15/00'),
        ('F21H 3/00', 'auto_replaced', 'result', 'F21V 3/00'),
        ('F24J 2/00', 'review_deprecated', None, None),
        ('F21M 3/00', 'review_deprecated', None, None),
        ('H01L 21/00', 'review_scope', None, None),
        ('H01L 33/00', 'review_scope', None, None),
        ('A01B 1/00', 'unchanged', None, None),
        ('G06F 3/01', 'unchanged', None, None),
        ('C13C3/02', 'auto_replaced', 'result', 'C13B 3/02'),
        ('H01L 21/00(2006.01)', 'review_scope', None, None),
    ]

    passed = 0
    failed = 0
    for code, exp_action, exp_key, exp_val in tests:
        result = rc.classify_code(code)
        ok = result['action'] == exp_action
        if ok and exp_key and exp_val:
            ok = result.get(exp_key) == exp_val
        if ok:
            passed += 1
        else:
            failed += 1
        detail = result.get('result', result.get('candidates', result.get('possible_new', '')))
        print(f"  {'✅' if ok else '❌'} {code:<16} → {result['action']:<20} {detail}")

    total = passed + failed + len(parse_tests) + len(field_tests) + len(date_tests) + len(ver_tests)
    print(f"\n自測: 全部通過" if failed == 0 else f"\n自測: {failed} 個失敗")
    return failed == 0


# ── 主程式 ───────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='IPC 批次重分類標記工具 — 標記專利 IPC 是否涉及版本異動',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''範例:
  python3 scripts/reclassify.py patents.xlsx -c IPC
  python3 scripts/reclassify.py patents.xlsx -c IPC -d 申請日
  python3 scripts/reclassify.py data.csv -c "分類號" -d "申請日" -o result.csv
  python3 scripts/reclassify.py --self-test

版本推斷：
  指定 -d 申請日欄位後，系統會根據申請日推斷 IPC 版本，
  只標記在該版本之後發生的異動。支援西元 (20050101) 和民國 (0940101) 格式。
''')
    parser.add_argument('input', nargs='?', help='輸入檔案路徑 (xlsx/csv/tsv)')
    parser.add_argument('-c', '--ipc-col', default='IPC', help='IPC 欄位名稱 (預設: IPC)')
    parser.add_argument('-d', '--date-col', help='申請日欄位名稱（用於推斷 IPC 版本）')
    parser.add_argument('-o', '--output', help='輸出檔案路徑 (預設: 輸入檔名加 _reclassified)')
    parser.add_argument('-v', '--verbose', action='store_true', help='顯示處理進度')
    parser.add_argument('--self-test', action='store_true', help='執行內建自測')
    parser.add_argument('--data', help='ipc_data.json 路徑 (預設: public/ipc_data.json)')
    args = parser.parse_args()

    # 初始化分類器
    data_path = args.data if args.data else None
    rc = IPCReclassifier(data_path)

    if args.self_test:
        print(f"IPCReclassifier {rc._stats}")
        print("\n執行自測...\n")
        ok = self_test(rc)
        sys.exit(0 if ok else 1)

    if not args.input:
        parser.print_help()
        sys.exit(1)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"錯誤：找不到檔案 {input_path}", file=sys.stderr)
        sys.exit(1)

    fmt = detect_format(input_path)
    output_path = Path(args.output) if args.output else input_path.with_name(
        input_path.stem + '_reclassified' + input_path.suffix)

    print(f"讀取 {input_path} ({fmt})...")
    rows, headers = read_input(input_path, args.ipc_col, fmt, args.date_col)
    print(f"  {len(rows):,} 筆資料，IPC 欄位: {args.ipc_col}"
          + (f"，申請日欄位: {args.date_col}" if args.date_col else ""))

    # 處理每筆資料
    results_count = defaultdict(int)
    version_stats = defaultdict(int) if args.date_col else None

    for i, row in enumerate(rows):
        ipc_value = row.get(args.ipc_col, '')

        # 推斷 IPC 版本
        patent_version = None
        if args.date_col:
            date_value = row.get(args.date_col, '')
            year = parse_apply_date(date_value)
            if year:
                patent_version = year_to_ipc_version(year)
                row['ipc_version_inferred'] = patent_version
                version_stats[patent_version] += 1
            else:
                row['ipc_version_inferred'] = ''

        classifications = rc.classify_field(ipc_value, patent_version)

        # 取最重要的 action
        actions = [c['action'] for c in classifications]
        if 'review_scope' in actions:
            primary_action = 'review_scope'
        elif 'review_deprecated' in actions:
            primary_action = 'review_deprecated'
        elif 'auto_replaced' in actions:
            primary_action = 'auto_replaced'
        else:
            primary_action = 'unchanged'

        row['reclassify_action'] = primary_action

        # result: 自動替換的結果
        replaced_codes = []
        for c in classifications:
            if c['action'] == 'auto_replaced':
                replaced_codes.append(c.get('result', c['original']))
            else:
                replaced_codes.append(c['original'])
        row['reclassify_result'] = '; '.join(replaced_codes) if 'auto_replaced' in actions else ''

        # candidates + detail
        candidates = [c.get('candidates', '') for c in classifications if c.get('candidates')]
        details = [c.get('detail', '') for c in classifications if c.get('detail')]
        possible = [c.get('possible_new', '') for c in classifications if c.get('possible_new')]

        row['reclassify_candidates'] = '; '.join(filter(None, candidates + possible))
        row['reclassify_detail'] = '; '.join(filter(None, details))

        results_count[primary_action] += 1

        if args.verbose and (i + 1) % 10000 == 0:
            print(f"  已處理 {i + 1:,} 筆...")

    # 寫出
    print(f"\n寫入 {output_path}...")
    write_output(output_path, rows, headers, fmt)

    # 摘要
    summary = generate_summary(results_count, len(rows), version_stats)
    print(f"\n{summary}")


if __name__ == '__main__':
    main()
