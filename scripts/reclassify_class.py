#!/usr/bin/env python3
"""
IPC 三階（Class 級）批次重分類標記工具

判斷每筆專利的 IPC class（三碼，如 H01、B23）是否有版本異動，
並列出涉及的版本轉換。

用法:
    python3 scripts/reclassify_class.py input.xlsx -c IPC
    python3 scripts/reclassify_class.py data.csv -c IPC -d 申請日 --verbose
    python3 scripts/reclassify_class.py --self-test
"""

from __future__ import annotations
import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

# ── 複用 reclassify.py 的日期解析 ────────────────────────────

IPC_VERSION_BY_YEAR = {
    **{y: '1994.01' for y in range(1994, 2000)},
    **{y: '2000.01' for y in range(2000, 2006)},
    **{y: f'{y}.01' for y in range(2006, 2030)},
}


def year_to_ipc_version(year: int) -> str:
    if year < 1994:
        return '1994.01'
    return IPC_VERSION_BY_YEAR.get(year, f'{year}.01')


def parse_apply_date(value) -> int | None:
    if not value:
        return None
    raw = re.sub(r'\D', '', str(value).strip())
    if not raw:
        return None
    try:
        n = int(raw)
    except ValueError:
        return None
    if n > 19000000:
        return n // 10000
    elif n > 800000:
        return n // 10000 + 1911
    elif n > 1900:
        return n
    elif n > 80:
        return n + 1911
    return None


def parse_ipc_field(value) -> list[str]:
    if not value or (isinstance(value, float) and str(value) == 'nan'):
        return []
    raw = str(value).strip()
    if not raw:
        return []
    parts = re.split(r'[,;|\n\r]+', raw)
    result = []
    for part in parts:
        part = part.strip().upper()
        # Remove version suffix
        part = re.sub(r'[\(\[]\d{4}[\.\-]?\d{0,2}[\)\]]?$', '', part).strip()
        # Remove TIPO prefix
        part = re.sub(r'^[A-Z]:\s*', '', part)
        # Extract IPC code
        m = re.match(r'([A-H]\d{2}[A-Z]?)(?:\s*\d+/\S+)?', part)
        if m:
            result.append(m.group(0).strip())
    return result


def extract_class(ipc_code: str) -> str | None:
    """從 IPC 代碼提取三碼 class（如 H01L 21/677 → H01）。"""
    code = ipc_code.strip().upper()
    code = re.sub(r'[\(\[]\d{4}[\.\-]?\d{0,2}[\)\]]?$', '', code).strip()
    code = re.sub(r'^[A-Z]:\s*', '', code)
    # Handle no-space: H01L21/677
    m = re.match(r'^([A-H]\d{2})', code)
    return m.group(1) if m else None


# ── Class 級分類器 ───────────────────────────────────────────

class IPCClassReclassifier:
    """IPC 三階（Class 級）重分類標記引擎。"""

    def __init__(self, data_path: str | Path | None = None):
        if data_path is None:
            data_path = Path(__file__).resolve().parent.parent / 'public' / 'ipc_data.json'
        with open(data_path, encoding='utf-8') as f:
            data = json.load(f)

        # 載入 subclass 名稱
        names_path = Path(data_path).parent / 'ipc_names.json'
        self.sub_names = {}
        if names_path.exists():
            with open(names_path, encoding='utf-8') as f:
                self.sub_names = json.load(f)

        si = data.get('subclass_index', {})
        dep_to = data.get('deprecated_to', {})
        dep_at = data.get('deprecated_at', {})
        intro = data.get('introduced_in', {})

        # 彙總每個 class 的異動版本
        self.class_versions = defaultdict(set)     # class → set of version strings
        self.class_deprecated = defaultdict(list)  # class → [(sub, dest, at)]
        self.class_introduced = defaultdict(list)  # class → [(sub, version)]
        self.class_changes = defaultdict(int)       # class → total change count

        for sub, info in si.items():
            cls = sub[:3]
            for d in info.get('donated', []):
                ver = d.get('version', '')
                if ver:
                    self.class_versions[cls].add(ver)
                    self.class_changes[cls] += 1
            for r in info.get('received', []):
                ver = r.get('version', '')
                if ver:
                    self.class_versions[cls].add(ver)

        for sub, dest in dep_to.items():
            cls = sub[:3]
            at = dep_at.get(sub, '?')
            dest_str = ', '.join(dest) if isinstance(dest, list) else dest
            self.class_deprecated[cls].append((sub, dest_str, at))

        for sub, ver in intro.items():
            cls = sub[:3]
            self.class_introduced[cls].append((sub, ver))

        # 所有已知 class
        self.all_classes = set()
        for sub in si:
            self.all_classes.add(sub[:3])
        for sub in dep_to:
            self.all_classes.add(sub[:3])
        for sub in intro:
            self.all_classes.add(sub[:3])

        affected = len([c for c in self.all_classes if c in self.class_versions])
        self._stats = f"loaded: {len(self.all_classes)} classes, {affected} affected"

    @staticmethod
    def _version_from_key(version_str: str) -> str:
        if '→' in version_str:
            return version_str.split('→')[0].strip()
        return version_str

    def get_class_description(self, cls: str) -> str:
        """取得 class 的中文描述（從 subclass 名稱推斷）。"""
        # 找這個 class 下的第一個 subclass 名稱
        for sub_code in sorted(self.sub_names.keys()):
            if sub_code.startswith(cls):
                name = self.sub_names[sub_code]
                if isinstance(name, str):
                    return name[:30]
        return ''

    def classify_class(self, cls: str, patent_version: str | None = None) -> dict:
        """對三碼 class 進行異動判斷。"""
        cls = cls.strip().upper()[:3]

        if cls not in self.class_versions and cls not in self.class_deprecated:
            desc = self.get_class_description(cls)
            return {
                'action': 'unchanged',
                'class': cls,
                'description': desc,
                'versions': '',
                'detail': '',
            }

        # 收集涉及的版本（版本感知篩選）
        relevant_versions = []
        for ver in sorted(self.class_versions.get(cls, [])):
            if patent_version:
                ver_from = self._version_from_key(ver)
                if ver_from < patent_version:
                    continue  # 異動在專利版本之前，已完成
            relevant_versions.append(ver)

        # 檢查廢棄
        dep_info = self.class_deprecated.get(cls, [])

        desc = self.get_class_description(cls)

        if not relevant_versions and not dep_info:
            return {
                'action': 'unchanged',
                'class': cls,
                'description': desc,
                'versions': '',
                'detail': '',
            }

        ver_str = ', '.join(relevant_versions) if relevant_versions else ''
        detail_parts = []
        if dep_info:
            for sub, dest, at in dep_info:
                sub_name = self.sub_names.get(sub, '')
                if isinstance(sub_name, str):
                    sub_name = sub_name[:20]
                dest_name = ''
                dest_first = dest.split(',')[0].strip()
                if dest_first in self.sub_names:
                    dest_name = self.sub_names[dest_first]
                    if isinstance(dest_name, str):
                        dest_name = dest_name[:20]
                detail_parts.append(f"{sub}({sub_name})→{dest}({dest_name}) [{at}]")
        if relevant_versions:
            detail_parts.append(f"涉及 {len(relevant_versions)} 個版本轉換")

        return {
            'action': 'has_changes',
            'class': cls,
            'description': desc,
            'versions': ver_str,
            'detail': '; '.join(detail_parts),
            'version_count': len(relevant_versions),
        }

    def classify_ipc_field(self, value, patent_version: str | None = None) -> list[dict]:
        """從 IPC 欄位提取 class 並分類。"""
        codes = parse_ipc_field(value)
        if not codes:
            return [{'action': 'unchanged', 'class': '', 'description': '', 'versions': '', 'detail': '空值'}]

        seen_classes = set()
        results = []
        for code in codes:
            cls = extract_class(code)
            if cls and cls not in seen_classes:
                seen_classes.add(cls)
                results.append(self.classify_class(cls, patent_version))

        return results if results else [{'action': 'unchanged', 'class': '', 'description': '', 'versions': '', 'detail': '無法解析'}]

    def get_all_classes_summary(self) -> list[dict]:
        """取得所有 class 的異動摘要（給網頁用）。"""
        result = []
        for cls in sorted(self.all_classes):
            info = self.classify_class(cls)
            result.append(info)
        return result


# ── I/O ──────────────────────────────────────────────────────

def detect_format(filepath: Path) -> str:
    ext = filepath.suffix.lower()
    if ext == '.xlsx': return 'xlsx'
    if ext == '.tsv': return 'tsv'
    if ext == '.csv': return 'csv'
    with open(filepath, encoding='utf-8') as f:
        if '\t' in f.readline(): return 'tsv'
    return 'csv'


def read_input(filepath, ipc_col, fmt, date_col=None):
    required = [ipc_col] + ([date_col] if date_col else [])
    if fmt == 'xlsx':
        try:
            import openpyxl
        except ImportError:
            print("錯誤：需要 openpyxl。安裝：pip install openpyxl", file=sys.stderr)
            sys.exit(1)
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h or '') for h in next(rows_iter)]
        for col in required:
            if col not in headers:
                print(f"錯誤：找不到欄位 '{col}'，可用: {headers}", file=sys.stderr)
                sys.exit(1)
        rows = [dict(zip(headers, [str(v) if v is not None else '' for v in row])) for row in rows_iter]
        wb.close()
        return rows, headers
    else:
        delim = '\t' if fmt == 'tsv' else ','
        with open(filepath, encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=delim)
            headers = reader.fieldnames or []
            for col in required:
                if col not in headers:
                    print(f"錯誤：找不到欄位 '{col}'，可用: {headers}", file=sys.stderr)
                    sys.exit(1)
            return list(reader), headers


def write_output(filepath, rows, headers, fmt):
    new_cols = ['ipc_class', 'class_action', 'class_description', 'affected_versions', 'class_detail', 'ipc_version_inferred']
    all_headers = headers + [c for c in new_cols if c not in headers]
    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(all_headers)
        for row in rows:
            ws.append([row.get(h, '') for h in all_headers])
        wb.save(filepath)
    else:
        delim = '\t' if fmt == 'tsv' else ','
        with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=all_headers, delimiter=delim, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)


# ── 自測 ─────────────────────────────────────────────────────

def self_test(rc: IPCClassReclassifier):
    print("  Class 分類測試：")
    tests = [
        ('H01', 'has_changes'),   # H01L 有大量異動
        ('G06', 'has_changes'),   # G06C/D/E/G/J 已廢棄
        ('A01', 'has_changes'),   # A01K 有異動
        ('A22', 'unchanged'),     # 屠宰，沒有異動
        ('D05', 'unchanged'),     # 縫紉，沒有異動
    ]
    passed = 0
    for cls, exp in tests:
        r = rc.classify_class(cls)
        ok = r['action'] == exp
        desc = r.get('description', '')[:20]
        ver = r.get('versions', '')[:40]
        print(f"    {'✅' if ok else '❌'} {cls} ({desc}) → {r['action']:<15} {ver}")
        if ok: passed += 1

    print("\n  IPC 欄位解析測試：")
    field_tests = [
        ('H01L 21/677', ['H01']),
        ('H01L 21/677; G06F 3/01', ['H01', 'G06']),
        ('B23K 26/00, A01B 1/00', ['B23', 'A01']),
    ]
    for raw, exp_classes in field_tests:
        results = rc.classify_ipc_field(raw)
        classes = [r['class'] for r in results]
        ok = classes == exp_classes
        print(f"    {'✅' if ok else '❌'} {raw:<30} → {classes}")
        if ok: passed += 1

    print("\n  版本感知測試：")
    # H01 在 1995→2000 有異動。若專利版本是 2010，應該只列 2010 之後的
    r1 = rc.classify_class('H01', '2023.01')
    r2 = rc.classify_class('H01', None)
    ok1 = r1['action'] == 'has_changes'
    ok2 = r2['action'] == 'has_changes'
    v1_count = r1.get('version_count', 0)
    v2_count = r2.get('version_count', 0)
    ok3 = v1_count < v2_count  # 有版本限制時應該少一些
    print(f"    {'✅' if ok1 else '❌'} H01 ver=2023.01 → {r1['action']} ({v1_count} versions)")
    print(f"    {'✅' if ok2 else '❌'} H01 ver=None    → {r2['action']} ({v2_count} versions)")
    print(f"    {'✅' if ok3 else '❌'} 版本限制後版本數較少: {v1_count} < {v2_count}")

    total = passed + 3
    print(f"\n自測: 全部通過" if (passed == len(tests) + len(field_tests)) and ok1 and ok2 and ok3 else f"\n自測: 有失敗")
    return True


# ── 主程式 ───────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='IPC 三階（Class 級）批次重分類標記工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''範例:
  python3 scripts/reclassify_class.py patents.xlsx -c IPC
  python3 scripts/reclassify_class.py data.csv -c IPC -d 申請日
  python3 scripts/reclassify_class.py --self-test
''')
    parser.add_argument('input', nargs='?', help='輸入檔案路徑')
    parser.add_argument('-c', '--ipc-col', default='IPC', help='IPC 欄位名稱')
    parser.add_argument('-d', '--date-col', help='申請日欄位名稱')
    parser.add_argument('-o', '--output', help='輸出檔案路徑')
    parser.add_argument('-v', '--verbose', action='store_true')
    parser.add_argument('--self-test', action='store_true')
    parser.add_argument('--data', help='ipc_data.json 路徑')
    args = parser.parse_args()

    rc = IPCClassReclassifier(args.data)

    if args.self_test:
        print(f"IPCClassReclassifier {rc._stats}")
        print("\n執行自測...\n")
        self_test(rc)
        sys.exit(0)

    if not args.input:
        parser.print_help()
        sys.exit(1)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"錯誤：找不到 {input_path}", file=sys.stderr)
        sys.exit(1)

    fmt = detect_format(input_path)
    output_path = Path(args.output) if args.output else input_path.with_name(
        input_path.stem + '_class_reclassified' + input_path.suffix)

    print(f"讀取 {input_path} ({fmt})...")
    rows, headers = read_input(input_path, args.ipc_col, fmt, args.date_col)
    print(f"  {len(rows):,} 筆，IPC: {args.ipc_col}" + (f"，申請日: {args.date_col}" if args.date_col else ""))

    counts = defaultdict(int)
    for i, row in enumerate(rows):
        ipc_value = row.get(args.ipc_col, '')

        patent_version = None
        if args.date_col:
            year = parse_apply_date(row.get(args.date_col, ''))
            if year:
                patent_version = year_to_ipc_version(year)
                row['ipc_version_inferred'] = patent_version

        results = rc.classify_ipc_field(ipc_value, patent_version)

        # 取最重要的 action
        has_changes = any(r['action'] == 'has_changes' for r in results)
        primary = 'has_changes' if has_changes else 'unchanged'

        row['ipc_class'] = ', '.join(r['class'] for r in results if r['class'])
        row['class_action'] = primary
        row['class_description'] = '; '.join(f"{r['class']}({r['description']})" for r in results if r['description'])
        row['affected_versions'] = '; '.join(r['versions'] for r in results if r['versions'])
        row['class_detail'] = '; '.join(r['detail'] for r in results if r['detail'])

        counts[primary] += 1
        if args.verbose and (i + 1) % 10000 == 0:
            print(f"  已處理 {i + 1:,} 筆...")

    print(f"\n寫入 {output_path}...")
    write_output(output_path, rows, headers, fmt)

    total = len(rows)
    print(f"\n{'='*50}")
    print(f"IPC 三階重分類標記報告")
    print(f"{'='*50}")
    print(f"總筆數: {total:,}")
    for action, label in [('unchanged', '無異動'), ('has_changes', '有異動')]:
        c = counts.get(action, 0)
        print(f"  {action:<15} {c:>8,} ({c/total*100:>5.1f}%) — {label}")


if __name__ == '__main__':
    main()
